import xlrd
from openpyxl import load_workbook
from openpyxl import worksheet
import re
import openpyxl
import openpyxl.styles as sty
k=0
list1=[]
list2=[]
list3=[]
def changexlmd(file1,sheet1_name,file2,sheet2_name):

    wb1 = load_workbook(file1)
    wb2 = load_workbook(file2)

    sheet1 = wb1[sheet1_name]
    sheet2 = wb2[sheet2_name]
    for i in range(1,sheet1.max_row):
        if type(sheet1.cell(row=i,column=7).value) == str:
            if " " in sheet1.cell(row=i, column=7).value:
                list1.append(sheet1.cell(row=i, column=7).value.split(" ")[0])
                list1.append(sheet1.cell(row=i, column=7).value.split(" ")[1])
            else:
                list1.append(sheet1.cell(row=i, column=7).value)





    print(list1)
    print(len(list1))

    for i in range(1,sheet2.max_row):
        if type(sheet2.cell(row=i,column=1).value) == str:
            if (sheet2.cell(row=i, column=9).value) == None:
                for w in range(0,len(list1)):
                    if (sheet2.cell(row=i, column=1).value) == list1[w]:
                        list2.append(sheet2.cell(row=i, column=1).value)
                        sheet2.cell(row=i, column=9).value = "In use - correct owner"

    print(list2)
    print(len(list2))

    wb2.save("/home/CORPUSERS/xp023799/Downloads/9999999.xlsx")

    # for i in range(1,sheet2.max_row):
    #     if type(sheet2.cell(row=i,column=21).value) == int:
    #        for b in range(len(list1)):
    #            if list1[b] == sheet2.cell(row=i,column=1).value:

    #
    # wb2.save("/home/CORPUSERS/xp023799/Downloads/ok.xlsx")

    # for i in range(1,sheet.max_row):
    #     if sheet.cell(row=i, column=3).value != None:
    #         if type(sheet.cell(row=i, column=3).value) == str:
    #             if ' ' not in sheet.cell(row=i, column=3).value and sheet.cell(row=i, column=3).value != '\n':
    #                 if sheet.cell(row=i, column=3).value.startswith('44'):
    #                     sheet.cell(row=i, column=3).value = sheet.cell(row=i, column=3).value.zfill(15)
    #                 elif '-' in sheet.cell(row=i, column=3).value:
    #                     sheet.cell(row=i, column=3).value = sheet.cell(row=i, column=3).value.replace('-', '')
    #                 else:
    #                     sheet.cell(row=i, column=3).value = sheet.cell(row=i, column=3).value
    #         else:
    #             if str(sheet.cell(row=i, column=3).value).startswith('44'):
    #                 sheet.cell(row=i, column=3).value = str(sheet.cell(row=i, column=3).value).zfill(15)
    #             else:
    #                 sheet.cell(row=i, column=3).value = str(sheet.cell(row=i, column=3).value)








            # if sheet1.cell(row,11).value == 'Dalian':
            #     o = sheet1.cell(row, 2).value
            #     if type(o) == float:
            #         if o > 9999:
            #             list1.append(int(o))
            #     else:
            #         if o != '' and o.isdigit():
            #             list1.append(int(o))


    # list2 = []
    # for row in range(sheet2.nrows):
    #
    #     if sheet2.cell(row,2).value == 'sony' or sheet2.cell(row,2).value =='Sony' or sheet2.cell(row,2).value == 'SONY' :
    #         k = sheet2.cell(row, 6).value
    #         if  type(k) == float:
    #             if k > 9999:
    #                 list2.append(int(k))
    #         else:
    #             if k !='' and k.isdigit():
    #                 list2.append(int(k))
    #
    # list3 = []
    # kk = (list(set(list1).intersection(list2)))
    #
    # for i in range(len(kk)):
    #     q = kk[i]
    #     list3.append(str(q))
    #
    # for qq in range(2,sheet.max_row):
    #     for j in range(len(list3)):
    #         if str(sheet.cell(row=qq, column=3).value)==list3[j]:
    #             for ww in range(1,sheet.max_column):
    #                 sheet.cell(row=qq, column=ww).fill = sty.PatternFill(fill_type='solid', fgColor="F0F8FF")





    # for row in range(sheet1.nrows):
    #     if sheet1.cell(row,5).value == 'SoMC' or sheet1.cell(row,5).value =='SOMC' or sheet1.cell(row,5).value == 'SONY':
    #         if sheet1.cell(row,11).value == 'Dalian':
    #             o = sheet1.cell(row, 2).value
    #             if type(o) == float:
    #                 if o > 9999:
    #                     list1.append(int(o))
    #             else:
    #                 if o != '' and o.isdigit():
    #                     list1.append(int(o))
    #
    # print(list1)
    # list2 = []
    # for row in range(sheet2.nrows):
    #
    #     if sheet2.cell(row,2).value == 'SOMC' or sheet2.cell(row,2).value =='SoMC':
    #         k = sheet2.cell(row, 6).value
    #         if  type(k) == float:
    #             if k > 9999:
    #                 list2.append(int(k))
    #         else:
    #             if k !='' and k.isdigit():
    #                 list2.append(int(k))
    #
    #
    # list3 = []
    # kk = (list(set(list1).intersection(list2)))
    #
    # for i in range(len(kk)):
    #     q = kk[i]
    #     list3.append(str(q))
    # print(list3)
    # for qq in range(2, sheet.max_row):
    #     for j in range(len(list3)):
    #         o = str(sheet.cell(row=qq, column=20).value)
    #         if type(o) == float:
    #
    #             k = int(o)
    #         else:
    #             if o != '' and type(o) != 'NoneType':
    #                 if x not in o and ' ' not in o:
    #                     if '-' in o:
    #                         k = int(o.replace("-", ""))
    #
    #                     else:
    #                         k = int(o)
    #
    #         if k == list3[j]:
    #             for ww in range(1, sheet.max_column):
    #                 sheet.cell(row=qq, column=ww).fill = sty.PatternFill(fill_type='solid', fgColor="FFB6C1")

    # wb.save("/home/CORPUSERS/xp023799/Downloads/ok.xlsx")







if __name__ == '__main__':
    changexlmd("/home/CORPUSERS/xp023799/Downloads/88888888888888888.xlsx",
               "Sheet1",
               "/home/CORPUSERS/xp023799/Downloads/hosi.xlsx",
               "IOP Samples",)

