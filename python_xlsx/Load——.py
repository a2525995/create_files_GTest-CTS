import xlrd
from openpyxl import load_workbook
import re
import openpyxl
import openpyxl.styles as sty
k=0


def changexlmd(file1,sheet1_name,file2,sheet2_name):
    list1 = []
    wb = load_workbook(file1)
    ll=0
    sheet = wb[sheet1_name]
    x='\n'

    sheet1 = xlrd.open_workbook(file1).sheet_by_name(sheet1_name)
    sheet2 = xlrd.open_workbook(file2).sheet_by_name(sheet2_name)


    for row in range(sheet1.nrows):
        if re.compile(r'Smartphone & Tablet', re.I).search(sheet1.cell(row,3).value) and re.compile(r'dalian', re.I).search(sheet1.cell(row,11).value):
            if re.compile(r'somc', re.I).search(sheet1.cell(row,5).value):

                o = sheet1.cell(row, 19).value
                if type(o) == float:

                    list1.append(int(o))
                else:
                    if o != '':
                        if  x not in o and ' ' not in o:
                            if '-' in o:
                                list1.append(int(o.replace("-", "")))

                            else:
                                list1.append(int(o))

    print(list1)
    list2 = []
    for row in range(sheet2.nrows):

        if re.compile(r'somc', re.I).search(sheet2.cell(row,5).value) or re.compile(r'sony', re.I).search(sheet2.cell(row,5).value):
            k = sheet2.cell(row, 11).value
            if type(k) == float:

                list2.append(int(k))
            else:
                if k != '' and 'ID' not in k :
                    if x not in k and ' ' not in k:
                        if '-' in k:
                            list2.append(int(k.replace("-", "")))

                        else:
                            list2.append(int(k))
    print(len(list2))

    list3 = []
    kk = (list(set(list1).intersection(list2)))
    list4 = []
    for i in range(len(kk)):
        q = kk[i]
        list3.append(str(q))
    print(len(list3))

    for qq in range(2, sheet.max_row):
        for j in range(len(list3)):
            o = sheet.cell(row=qq, column=20).value
            if type(o) == int:
                k = str(o)
            else:
                if type(o)==str:
                    if '-' in o:
                        k = str(o.replace("-", ""))

                    else:
                         k = str(o)

            if k == list3[j]:
                list4.append(k)
    print(type(sheet.cell(row=1994,column=20).value))
    print(list4)
    print(list(set(list3).difference(set(list4))))
    wb.save("/home/CORPUSERS/xp023799/Downloads/4444.xlsx")







if __name__ == '__main__':
    changexlmd("/home/CORPUSERS/xp023799/Downloads/1111.xlsx","PIOT Devices List",
               "/home/CORPUSERS/xp023799/Downloads/DALIAN+SoMC_Device_List.xlsx","SOMC Phone List_Prototype")

